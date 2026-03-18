"""
Excel レポート生成ヘルパー

使い方:
  from excel_helpers import ReportWorkbook
  wb = ReportWorkbook()
  ws = wb.add_sheet("業務一覧")
  wb.set_header(ws, 1, ["#", "業務名", "頻度"], [5, 25, 10])
  wb.set_row(ws, 2, [1, "記事編集", "日次"])
  wb.set_input(ws, 3, 2, 30, "1回あたりの時間（分）を変更できます")
  wb.set_formula(ws, 3, 4, "=B3*C3/60")
  wb.save("output.xlsx")
"""
import sys
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter


# ===== カラーパレット（HTMLレポートと統一） =====
HEADER_BG = "2C2C34"       # charcoal
HEADER_FG = "FFFFFF"
ACCENT = "E07A5F"           # peach
SAGE = "6B8F71"             # sage
PLUM = "8A6C8A"             # plum
INPUT_BG = "FFF8E1"         # 薄い黄色（入力セル）
BORDER_COLOR = "E8E2DA"
LIGHT_PEACH = "FDF0EB"      # 1位の行背景
LIGHT_SAGE = "EEF4EF"       # 2位の行背景
LIGHT_PLUM = "F3EEF3"       # 3位の行背景

# XLSXではWebフォント（Noto Serif JP等）が使えないためYu Gothicを使用
FONT_NAME = "Yu Gothic"

# ===== スタイル定義 =====
_header_font = Font(name=FONT_NAME, bold=True, color=HEADER_FG, size=10)
_header_fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
_body_font = Font(name=FONT_NAME, size=10)
_bold_font = Font(name=FONT_NAME, size=10, bold=True)
_input_fill = PatternFill(start_color=INPUT_BG, end_color=INPUT_BG, fill_type="solid")
_thin_border = Border(bottom=Side(style="thin", color=BORDER_COLOR))
_header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
_wrap_align = Alignment(vertical="top", wrap_text=True)

# 順位別
RANK_FILLS = [
    PatternFill(start_color=LIGHT_PEACH, end_color=LIGHT_PEACH, fill_type="solid"),
    PatternFill(start_color=LIGHT_SAGE, end_color=LIGHT_SAGE, fill_type="solid"),
    PatternFill(start_color=LIGHT_PLUM, end_color=LIGHT_PLUM, fill_type="solid"),
]
RANK_FONTS = [
    Font(name=FONT_NAME, size=10, bold=True, color=ACCENT),
    Font(name=FONT_NAME, size=10, bold=True, color=SAGE),
    Font(name=FONT_NAME, size=10, bold=True, color=PLUM),
]


class ReportWorkbook:
    """Excel レポート生成のラッパー"""

    def __init__(self):
        self.wb = Workbook()
        # デフォルトシートは最初の add_sheet で上書きされる
        self._first = True

    def add_sheet(self, name: str):
        """シートを追加して返す。最初の呼び出しはデフォルトシートをリネーム"""
        if self._first:
            ws = self.wb.active
            ws.title = name
            self._first = False
        else:
            ws = self.wb.create_sheet(name)
        return ws

    def set_header(self, ws, row: int, headers: list[str], widths: list[int] = None):
        """ヘッダー行を設定"""
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=ci, value=h)
            c.font = _header_font
            c.fill = _header_fill
            c.alignment = _header_align
            c.border = _thin_border
        if widths:
            for ci, w in enumerate(widths, 1):
                ws.column_dimensions[get_column_letter(ci)].width = w

    def set_row(self, ws, row: int, values: list, font=None, fill=None):
        """データ行を設定"""
        for ci, v in enumerate(values, 1):
            c = ws.cell(row=row, column=ci, value=v)
            c.font = font or _body_font
            c.alignment = _wrap_align
            c.border = _thin_border
            if fill:
                c.fill = fill

    def set_bold_row(self, ws, row: int, values: list, fill=None):
        """太字のデータ行"""
        self.set_row(ws, row, values, font=_bold_font, fill=fill)

    def set_rank_row(self, ws, row: int, values: list, rank: int):
        """順位付き行（1位=peach, 2位=sage, 3位=plum）"""
        idx = min(rank - 1, 2)
        self.set_row(ws, row, values, fill=RANK_FILLS[idx])
        ws.cell(row=row, column=1).font = RANK_FONTS[idx]

    def set_input(self, ws, row: int, col: int, value, comment_text: str = None):
        """入力セル（黄色背景）"""
        c = ws.cell(row=row, column=col, value=value)
        c.fill = _input_fill
        c.font = _body_font
        c.border = _thin_border
        if comment_text:
            c.comment = Comment(comment_text, "システム")

    def set_formula(self, ws, row: int, col: int, formula: str, fmt: str = "0.0", font=None):
        """数式セル"""
        c = ws.cell(row=row, column=col, value=formula)
        c.font = font or _body_font
        c.number_format = fmt
        c.border = _thin_border

    def set_section_title(self, ws, row: int, col: int, text: str, font=None):
        """セクションタイトル（太字・色付き）"""
        c = ws.cell(row=row, column=col, value=text)
        c.font = font or Font(name=FONT_NAME, size=12, bold=True)

    def set_subtitle(self, ws, row: int, col: int, text: str):
        """説明文（グレー小文字）"""
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name=FONT_NAME, size=9, color="888888")

    def accent_font(self):
        return RANK_FONTS[0]

    def sage_font(self):
        return RANK_FONTS[1]

    def plum_font(self):
        return RANK_FONTS[2]

    def total_font(self):
        return Font(name=FONT_NAME, size=12, bold=True, color=ACCENT)

    def save(self, path: str):
        """ファイル保存"""
        self.wb.save(path)
        print(f"OK: {path}")
